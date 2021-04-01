# TO DO?
# - Aller chercher les uuid des responsables de collection
# - Créer un cache d'uuid pour les auteurs d'ouvrages et les lier à des E21
# - Dater un F27 : time-span + propriété cidoc ou dcterms:date ?

import argparse
from openpyxl import load_workbook
from rdflib import Graph, Literal, Namespace, DCTERMS, RDF, RDFS, SKOS, URIRef, XSD, URIRef as u, Literal as l
import sys
from sherlockcachemanagement import Cache

# Arguments
parser = argparse.ArgumentParser()
parser.add_argument("--iiif_excel_coll")
parser.add_argument("--iiif_excel_index")
parser.add_argument("--output_ttl")
parser.add_argument("--cache_40CM")
args = parser.parse_args()

# Cache
cache_40CM = Cache(args.cache_40CM)

# Initialisation du graphe
output_graph = Graph()

# Namespaces
crm_ns = Namespace("http://www.cidoc-crm.org/cidoc-crm/")
crmdig_ns = Namespace("http://www.ics.forth.gr/isl/CRMdig/")
iremus_ns = Namespace("http://data-iremus.huma-num.fr/id/")
lrmoo_ns = Namespace("http://www.cidoc-crm.org/lrmoo/")
sdt_ns = Namespace("http://data-iremus.huma-num.fr/datatypes/")
she_ns = Namespace("http://data-iremus.huma-num.fr/ns/sherlock#")

output_graph.bind("crm", crm_ns)
output_graph.bind("dcterms", DCTERMS)
output_graph.bind("lrm", lrmoo_ns)
output_graph.bind("sdt", sdt_ns)
output_graph.bind("skos", SKOS)
output_graph.bind("she", she_ns)
output_graph.bind("crmdig", crmdig_ns)

# Fonctions
a = RDF.type


def crm(x):
    return URIRef(crm_ns[x])


def crmdig(x):
    return URIRef(crmdig_ns[x])


def lrm(x):
    return URIRef(lrmoo_ns[x])


def she(x):
    return URIRef(iremus_ns[x])


def t(s, p, o):
    output_graph.add((s, p, o))


# Fichiers Excel
# Index des collections
wb_index = load_workbook(args.iiif_excel_index)
index = wb_index.active
# Images de la collection
wb_img = load_workbook(args.iiif_excel_coll)
img = wb_img.active

sys.exit()

#####################################################################
# LA COLLECTION
#####################################################################

collection_id = img.cell_value(4, 0)

collection = she(cache_40CM.get_uuid(["collection", "uuid"], True))
t(collection, a, crmdig("D1_Digital_Object"))
# Appellation
collection_E41 = she(cache_40CM.get_uuid(["collection", "E41"], True))
t(collection_E41, a, crm("E41_Appellation"))
t(collection, crm("P1_is_identified_by"), collection_E41)
t(collection_E41, RDFS.label, Literal(index.cell_value(4, 1)))
t(collection, crm("P2_has_type"), she("14926d58-83e7-4414-90a8-1a3f5ca8fec1"))
# Creation
collection_E65 = she(cache_40CM.get_uuid(["collection", "E65"], True))
t(collection_E65, a, crm("E65_Creation"))
t(collection_E65, crm("P94_has_created"), collection)
t(collection_E65, crm("P14_carried_out_by"), Literal(index.cell_value(4, 2)))  # Aller chercher l'uuid du-de la responsable
# Licence
collection_E30 = she(cache_40CM.get_uuid(["collection", "E30"], True))
t(collection_E30, a, crm("E30_Right"))
t(collection, crm("P104_is_subject_to"), collection_E30)
t(collection_E30, RDFS.label, Literal(index.cell_value(4, 5)))
# Attribution
t(collection, crm("P105_right_held_by"), she("48a8e9ad-4264-4b0b-a76d-953bc9a34498"))


#####################################################################
# 1. UNE PUBLICATION NUMERISEE
#####################################################################

if index.cell_value(4, 3) == "Livre":

    # Work
    livre_F1 = she(cache_40CM.get_uuid(["collection", "livre", "F1"], True))
    t(livre_F1, a, lrm("F1_Work"))
    livre_E41 = she(cache_40CM.get_uuid(["collection", "livre", "E41"], True))
    t(livre_F1, crm("P1_is_identified_by"), livre_E41)
    t(livre_F1, RDFS.label, Literal(index.cell_value(4, 6)))
    # Work Conception
    if index.cell_value(4, 7) or index.cell_value(4, 9) != None:
        livre_F27 = she(cache_40CM.get_uuid(["collection", "livre", "F27"], True))
        t(livre_F27, a, lrm("F27_Work_Conception"))
        t(livre_F27, lrm("R16_initiated"), livre_F1)
        if index.cell_value(4, 7) != None:
            t(livre_F27, crm("P14_carried_out_by"), Literal(index.cell_value(4, 7)))
        if index.cell_value(4, 9) != None:
            # E52 OU DCTERMS:DATE?
            # livre_E52 = she(cache_40CM.get_uuid(["collection", "livre", "E52"], True))
            # t(livre_F27, crm("P4_has_time-span"), livre_E52)
            # t(livre_E52, DCTERMS.date, Literal(index.cell_value(4, 9)))
            t(livre_F27, DCTERMS.date, Literal(index.cell_value(4, 9), datatype=XSD.date))

    # Expression
    livre_F2 = she(cache_40CM.get_uuid(["collection", "livre", "F2"], True))
    t(livre_F1, lrm("R3_is_realised_in"), livre_F2)
    t(livre_F2, a, lrm("F2_Expression"))
    # Expression Creation
    if index.cell_value(4, 8) or index.cell_value(4, 10) != None:
        livre_F28 = she(cache_40CM.get_uuid(["collection", "livre", "F28"], True))
        t(livre_F28, a, lrm("F28_Expression_Creation"))
        t(livre_F28, lrm("R17_created"), livre_F2)
        if index.cell_value(4, 8) != None:
            t(livre_F28, crm("P14_carried_out_by"), Literal(index.cell_value(4, 8)))
        if index.cell_value(4, 10) != None:
            t(livre_F28, DCTERMS.date, Literal(index.cell_value(4, 10), datatype=XSD.date))

    # Manifestation
    livre_F3 = she(cache_40CM.get_uuid(["collection", "livre", "F3"], True))
    t(livre_F3, a, lrm("F3_Manifestation"))
    t(livre_F3, lrm("R4_embodies"), livre_F2)
    # Item
    livre_F5 = she(cache_40CM.get_uuid(["collection", "livre", "F5"], True))
    t(livre_F5, a, lrm("F5_Item"))
    t(livre_F5, lrm("R7_is_materialization_of"), livre_F3)
    # AJOUT D'UN F32 OU REDONDANT AVEC LE F28?

    #####################################################################
    # LES PAGES DE LA PUBLICATION
    #####################################################################

    def id_page(row, column):
        id = img.cell_value(row, column)
        try:
            # La page comme support physique
            page_E18 = she(cache_40CM.get_uuid(["collection", "livre", id, "E18"], True))
            t(page_E18, a, crm("E18_Physical_Object"))
            t(livre_F5, crm("P46_is_composed_of"), page_E18)

            # La page comme support sémiotique
            page_E90 = she(cache_40CM.get_uuid(["collection", "livre", id, "E90"], True))
            t(page_E90, a, crm("E90_Symbolic_Object"))
            t(livre_F2, lrm("R15_has_fragment"), page_E90)
            t(page_E18, crm("P128_carries"), page_E90)

            # Identifiant
            page_E42 = she(cache_40CM.get_uuid(["collection", "livre", id, "E42"], True))
            t(page_E42, crm("P2_has_type"), she("466bb717-b90f-4104-8f4e-5a13fdde3bc3"))
            t(page_E90, crm("P1_is_identified_by"), page_E42)
            t(page_E42, RDFS.label, Literal(img.cell_value(row, column), datatype=XSD.integer))

            # Numérisation de la page
            page_D2 = she(cache_40CM.get_uuid(["collection", "livre", "D2"], True))
            t(page_D2, a, crmdig("D2_Digitization_Process"))
            t(page_D2, crmdig("L1_digitized"), page_E18)
            page_D1 = she(cache_40CM.get_uuid(["collection", "livre", id, "D1"], True))
            t((page_D1), a, crmdig("D1_Digital_Object"))
            t(page_D2, crmdig("L11_had_output"), page_D1)
            t(page_D1, crm("130_shows_features_of"), page_E90)
            t(collection, crm("P106_is_composed_of"), page_D1)

            # Transcription de la page TO DO

            id_page(row + 1, column)
        except:
            pass

    id_page(4, 2)


#####################################################################
# 2. DES IMAGES INDIVIDUELLES
#####################################################################

if index.cell_value(4, 3) == "Images":

    def id_img(row, column):
        sheet_img = wb_index.sheet_by_index(0)
        id = sheet_img.cell_value(row, column)
        try:
            # L'image comme support physique
            image_E22 = she(cache_40CM.get_uuid(["collection", id, "E22"], True))
            t(image_E22, a, crm("E22_Human-Made_Object"))

            # L'image comme support sémiotique
            image_E36 = she(cache_40CM.get_uuid(["collection", id, "E36"], True))
            t(image_E22, crm("P65_shows_visual_item"), )

            id_img(row + 1, column)
        except:
            pass

    id_img(4, 2)

output_graph.serialize(destination=args.output_ttl, format="turtle", base="http://data-iremus.huma-num.fr/id/")
cache_40CM.bye()
