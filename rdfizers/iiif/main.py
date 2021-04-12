# TODO Transformer les auteurs d'ouvrages en E21?

import argparse
from openpyxl import load_workbook
from rdflib import Graph, Literal, Namespace, DCTERMS, RDF, RDFS, SKOS, URIRef, XSD, URIRef as u, Literal as l
import sys
from sherlockcachemanagement import Cache

# Arguments
parser = argparse.ArgumentParser()
parser.add_argument("--collection_id")
parser.add_argument("--iiif_excel_coll")
parser.add_argument("--iiif_excel_index")
parser.add_argument("--output_ttl")
parser.add_argument("--cache_40CM")
args = parser.parse_args()

# Cache
cache_40CM = Cache(args.cache_40CM)

# Initialisation du graphe
output_graph = Graph()

# Namespaces
crm_ns = Namespace("http://www.cidoc-crm.org/cidoc-crm/")
crmdig_ns = Namespace("http://www.ics.forth.gr/isl/CRMdig/")
iremus_ns = Namespace("http://data-iremus.huma-num.fr/id/")
lrmoo_ns = Namespace("http://www.cidoc-crm.org/lrmoo/")
sdt_ns = Namespace("http://data-iremus.huma-num.fr/datatypes/")
she_ns = Namespace("http://data-iremus.huma-num.fr/ns/sherlock#")

output_graph.bind("crm", crm_ns)
output_graph.bind("dcterms", DCTERMS)
output_graph.bind("lrm", lrmoo_ns)
output_graph.bind("sdt", sdt_ns)
output_graph.bind("skos", SKOS)
output_graph.bind("she", she_ns)
output_graph.bind("crmdig", crmdig_ns)

# Fonctions
a = RDF.type


def crm(x):
    return URIRef(crm_ns[x])


def crmdig(x):
    return URIRef(crmdig_ns[x])


def lrm(x):
    return URIRef(lrmoo_ns[x])


def she(x):
    return URIRef(iremus_ns[x])


def t(s, p, o):
    output_graph.add((s, p, o))


# Fichiers Excel
# Index des collections
wb_index = load_workbook(args.iiif_excel_index)
index = wb_index.active
# Images de la collection
wb_img = load_workbook(args.iiif_excel_coll)
img = wb_img.active

#####################################################################
# LA COLLECTION
#####################################################################

collection_row = None

for row in index:
    if row[0].value == args.collection_id:
        collection_row = row
        break

collection = she(cache_40CM.get_uuid(["collection", "uuid"], True))
t(collection, a, crmdig("D1_Digital_Object"))
# Appellation
collection_E41 = she(cache_40CM.get_uuid(["collection", "E41"], True))
t(collection_E41, a, crm("E41_Appellation"))
t(collection, crm("P1_is_identified_by"), collection_E41)
t(collection_E41, RDFS.label, Literal(collection_row[1].value))
t(collection, crm("P2_has_type"), she("14926d58-83e7-4414-90a8-1a3f5ca8fec1"))
# Creation
collection_E65 = she(cache_40CM.get_uuid(["collection", "E65"], True))
t(collection_E65, a, crm("E65_Creation"))
t(collection_E65, crm("P94_has_created"), collection)
t(collection_E65, crm("P14_carried_out_by"), she(collection_row[2].value))
# Licence
collection_E30 = she(cache_40CM.get_uuid(["collection", "E30"], True))
t(collection_E30, a, crm("E30_Right"))
t(collection, crm("P104_is_subject_to"), collection_E30)
t(collection_E30, RDFS.label, Literal(collection_row[5].value))
# Attribution
t(collection, crm("P105_right_held_by"), she("48a8e9ad-4264-4b0b-a76d-953bc9a34498"))

#####################################################################
# 1. UNE PUBLICATION NUMERISEE
#####################################################################

if collection_row[3].value == "Edition":

    # Work
    livre_F1 = she(cache_40CM.get_uuid(["collection", "livre", "F1"], True))
    t(livre_F1, a, lrm("F1_Work"))
    livre_E35 = she(cache_40CM.get_uuid(["collection", "livre", "E41"], True))
    t(livre_F1, crm("P102_has_title"), livre_E35)
    t(livre_E35, a, crm("E35_Title"))
    t(livre_E35, RDFS.label, Literal(collection_row[1].value))

    # Work Conception
    livre_F27 = she(cache_40CM.get_uuid(["collection", "livre", "F27"], True))
    t(livre_F27, a, lrm("F27_Work_Conception"))
    t(livre_F27, lrm("R16_initiated"), livre_F1)
    t(livre_F27, crm("P14_carried_out_by"), Literal(collection_row[7].value))
    if collection_row[8].value != None:
        livre_F27_E52 = she(cache_40CM.get_uuid(["collection", "livre", "F27_E52"], True))
        t(livre_F27, crm("P4_has_time-span"), livre_F27_E52)
        t(livre_F27_E52, crm("P80_end_is_qualified_by"), Literal(collection_row[8].value))

    # Expression
    livre_F2 = she(cache_40CM.get_uuid(["collection", "livre", "F2"], True))
    t(livre_F1, lrm("R3_is_realised_in"), livre_F2)
    t(livre_F2, a, lrm("F2_Expression"))
    # Expression Creation
    livre_F28 = she(cache_40CM.get_uuid(["collection", "livre", "F28"], True))
    t(livre_F28, a, lrm("F28_Expression_Creation"))
    t(livre_F28, lrm("R17_created"), livre_F2)
    t(livre_F28, crm("P14_carried_out_by"), Literal(collection_row[7].value))
    if collection_row[9].value != None:
        livre_F28_E52 = she(cache_40CM.get_uuid(["collection", "livre", "F28_E52"], True))
        t(livre_F28, crm("P4_has_time-span"), livre_F28_E52)
        t(livre_F28_E52, crm("P80_end_is_qualified_by"), Literal(collection_row[9].value))

    # Manifestation
    livre_F3 = she(cache_40CM.get_uuid(["collection", "livre", "F3"], True))
    t(livre_F3, a, lrm("F3_Manifestation"))
    t(livre_F3, lrm("R4_embodies"), livre_F2)
    ## Manifestation Creation
    livre_F30 = she(cache_40CM.get_uuid(["collection", "livre", "F30"], True))
    t(livre_F30, a, lrm("F30_Manifestation_Creation"))
    t(livre_F30, lrm("R24_created"), livre_F3)
    t(livre_F30, crm("P92_brought_into_existence"), livre_F2)
    if collection_row[10].value != None:
        livre_F30_E52 = she(cache_40CM.get_uuid(["collection", "livre", "F30_E52"], True))
        t(livre_F30, crm("P4_has_time-span"), livre_F30_E52)
        t(livre_F30_E52, crm("P80_end_is_qualified_by"), Literal(collection_row[10].value))
    # Item
    livre_F5 = she(cache_40CM.get_uuid(["collection", "livre", "F5"], True))
    t(livre_F5, a, lrm("F5_Item"))
    t(livre_F5, lrm("R7_is_materialization_of"), livre_F3)

    #####################################################################
    # LES PAGES DE LA PUBLICATION
    #####################################################################

    img_row = None

    for row in img:
        if row[1].value == args.collection_id:
            img_row = row
            id = img_row[0].value

            #La page comme support physique
            page_E18 = she(cache_40CM.get_uuid(["collection", "livre", "pages", id, "E18"], True))
            t(page_E18, a, crm("E18_Physical_Object"))
            t(livre_F5, crm("P46_is_composed_of"), page_E18)

            # La page comme support sémiotique
            page_E90 = she(cache_40CM.get_uuid(["collection", "livre", "pages", id, "E90"], True))
            t(page_E90, a, crm("E90_Symbolic_Object"))
            t(livre_F2, lrm("R15_has_fragment"), page_E90)
            t(page_E18, crm("P128_carries"), page_E90)

            # Identifiant
            page_E42 = she(cache_40CM.get_uuid(["collection", "livre", "pages", id, "E42"], True))
            t(page_E90, crm("P1_is_identified_by"), page_E42)
            t(page_E42, a, crm("E42_Identifier"))
            t(page_E42, RDFS.label, Literal(id))

            #Numéro de la page
            page_E41 = she(cache_40CM.get_uuid(["collection", "livre", "pages", id, "E42_numéro"], True))
            t(page_E41, crm("P2_has_type"), she("466bb717-b90f-4104-8f4e-5a13fdde3bc3"))
            t(page_E90, crm("P1_is_identified_by"), page_E41)
            t(page_E41, a, crm("E41_Appellation"))
            t(page_E41, RDFS.label, Literal(f"Page {img_row[3].value}"))

            # Numérisation de la page
            page_D2 = she(cache_40CM.get_uuid(["collection", "livre", "pages", "D2"], True))
            t(page_D2, a, crmdig("D2_Digitization_Process"))
            t(page_D2, crmdig("L1_digitized"), page_E18)
            page_D1 = she(cache_40CM.get_uuid(["collection", "livre", "pages", id, "D1"], True))
            t((page_D1), a, crmdig("D1_Digital_Object"))
            t(page_D2, crmdig("L11_had_output"), page_D1)
            t(page_D1, crm("130_shows_features_of"), page_E90)
            t(collection, crm("P106_is_composed_of"), page_D1)

            # Transcription de la page TO DO

#####################################################################
# 2. UNE COLLECTION D'IMAGES INDIVIDUELLES
#####################################################################

if collection_row[3].value == "Images":

    img_row = None

    for row in img:
        # TODO Ajouter dans le tableau d'Anne une colonne "collection" à l'index [1]
        if row[1].value == args.collection_id:
            img_row = row
            id = img_row[0].value

            # L'image comme support physique
            image_E22 = she(cache_40CM.get_uuid(["collection", id, "E22"], True))
            t(image_E22, a, crm("E22_Human-Made_Object"))
            ## Création de l'image
            image_E22_E65 = she(cache_40CM.get_uuid(["collection", id, "E22_E65"], True))
            t(image_E22_E65, a, crm("E65_Creation"))

            # L'image comme support sémiotique
            image_E36 = she(cache_40CM.get_uuid(["collection", id, "E36"], True))
            t(image_E22, crm("P65_shows_visual_item"), image_E36)
            t(image_E36, a, crm("P1_has_identifier"), img_row[0])
            ## Titre sur l'image
            image_E36_E35 = she(cache_40CM.get_uuid(["collection", id, "E36_E35"], True))
            t(image_E36, crm("P102_has_title"), image_E36_E35)
            t(image_E36_E35, RDFS.label, img_row[2])
            t(image_E36_E35, crm("P2_has_type"), she("01a07474-f2b9-4afd-bb05-80842ecfb527"))
            # TODO Ajouter les autres types de titres

            # TODO Ajouter les E13 d'indexation



serialization = output_graph.serialize(format="turtle", base="http://data-iremus.huma-num.fr/id/")
with open(args.output_ttl, "wb") as f:
    f.write(serialization)
cache_40CM.bye()
