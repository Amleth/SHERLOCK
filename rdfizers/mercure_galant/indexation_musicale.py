import argparse
from openpyxl import load_workbook
import sys
# sys.path.insert(0, '/rdfizers/helpers')
# import helpers_rdf.py
# from ..rdfizers import *
# from helpers_rdf import init_graph
from sherlockcachemanagement import Cache
from rdflib import DCTERMS, Graph, Namespace, RDF, SKOS, URIRef as u, Literal as l

######################################################################################
# CACHES
######################################################################################

parser = argparse.ArgumentParser()
parser.add_argument("--xlsx")
parser.add_argument("--ttl")
parser.add_argument("--cache")
parser.add_argument("--cache_corpus")
args = parser.parse_args()

cache = Cache(args.cache)
cache_corpus = Cache(args.cache_corpus)

######################################################################################
# INITIALISATION DU GRAPHE
######################################################################################

g = Graph()

crm_ns = Namespace("http://www.cidoc-crm.org/cidoc-crm/")
crmdig_ns = Namespace("http://www.ics.forth.gr/isl/CRMdig/")
iremus_ns = Namespace("http://data-iremus.huma-num.fr/id/")
lrmoo_ns = Namespace("http://www.cidoc-crm.org/lrmoo/")
sdt_ns = Namespace("http://data-iremus.huma-num.fr/datatypes/")
sherlock_ns = Namespace("http://data-iremus.huma-num.fr/ns/sherlock#")

g.bind("crm", crm_ns)
g.bind("dcterms", DCTERMS)
g.bind("lrm", lrmoo_ns)
g.bind("sdt", sdt_ns)
g.bind("skos", SKOS)
g.bind("crmdig", crmdig_ns)
g.bind("she_ns", sherlock_ns)
g.bind("she", iremus_ns)

######################################################################################
# FONCTIONS RDF
######################################################################################

a = RDF.type

def crm(x):
    return crm_ns[x]

def crmdig(x):
    return crmdig_ns[x]

def lrm(x):
    return lrmoo_ns[x]

def she(x):
    return iremus_ns[x]

def she_ns(x):
    return sherlock_ns[x]

def t(s, p, o):
    g.add((s, p, o))


#######################################################################################
# RECUPERATION DES DONNEES
#######################################################################################

# Fichier Excel
sheet = load_workbook(args.xlsx).active

for row in sheet.iter_rows(min_row=2):

    if row[5].value:
        id = row[5].value.replace(" ", "_").lower()
    else:
        id = (row[2].value + row[4].value).lower().replace(" ", "_").replace(",", "_")

    # Oeuvre musicale (F1 et F2)
    F1_oeuvre_uuid = she(cache.get_uuid([id, "oeuvre musicale", "F1", "uuid"], True))
    F2_oeuvre_uuid = she(cache.get_uuid([id, "oeuvre musicale", "F2", "uuid"], True))
    t(F1_oeuvre_uuid, lrm("R3_is_realised_in"), F2_oeuvre_uuid)

    # Incipit de l'oeuvre musicale
    if row[5].value:
        F1_oeuvre_appellation = she(cache.get_uuid([id, "oeuvre musicale", "F1", "E41", "uuid"], True))
        t(F1_oeuvre_uuid, crm("P1_is_identified_by"), F1_oeuvre_appellation)
        t(F1_oeuvre_appellation, a, crm("E41_Appellation"))
        t(F1_oeuvre_appellation, a, crm("E33_Linguistic_Object"))
        ## E13 : l'appellation est de type "incipit"
        F1_oeuvre_appellation_E13 = she(cache.get_uuid([id, "oeuvre musicale", "F1", "E41", "E13", "type incipit"], True))
        t(F1_oeuvre_appellation_E13, a, crm("E13_Attribute_Assignement"))
        t(F1_oeuvre_appellation_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
        t(F1_oeuvre_appellation_E13, crm("P140_assigned_attribute_to"), F1_oeuvre_appellation)
        t(F1_oeuvre_appellation_E13, crm("P141_assigned"), she("e43ce57c-8bf7-43b5-87a2-cf8c140030a6"))
        t(F1_oeuvre_appellation_E13, crm("P177_assigned_property_type"), crm("P2_has_type"))
        ## E13 : l'appellation a pour contenu symbolique...
        F1_oeuvre_appellation_P190_E13 = she(cache.get_uuid([id, "oeuvre musicale", "F1", "E41", "E13", "P190"], True))
        t(F1_oeuvre_appellation_P190_E13, a, crm("E13_Attribute_Assignement"))
        t(F1_oeuvre_appellation_P190_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
        t(F1_oeuvre_appellation_P190_E13, crm("P140_assigned_attribute_to"), F1_oeuvre_appellation)
        t(F1_oeuvre_appellation_P190_E13, crm("P141_assigned"), l(row[5].value))
        t(F1_oeuvre_appellation_P190_E13, crm("P177_assigned_property_type"), crm("P190_has_symbolic_content"))
        ## E13 : code de l'incipit
        F1_oeuvre_appellation_P1_E13 = she(cache.get_uuid([id, "oeuvre musicale", "F1", "E41", "E13", "code incipit"], True))
        t(F1_oeuvre_appellation_P1_E13, a, crm("E13_Attribute_Assignement"))
        t(F1_oeuvre_appellation_P1_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
        t(F1_oeuvre_appellation_P1_E13, crm("P140_assigned_attribute_to"), F1_oeuvre_appellation)
        t(F1_oeuvre_appellation_P1_E13, crm("P141_assigned"), l(row[39].value))
        t(F1_oeuvre_appellation_P1_E13, crm("P177_assigned_property_type"), crm("P190_has_symbolic_content"))

    # L'oeuvre musicale et composée d'un texte
    # TODO Vérifier si l'on souhaite vraiment utiliser R5 (erreur dans la modélisation?)
    F1_texte_uuid = she(cache.get_uuid([id, "texte", "F1", "uuid"], True))
    F2_texte_uuid = she(cache.get_uuid([id, "texte", "F2", "uuid"], True))
    t(F1_texte_uuid, lrm("R3_is_realised_in"), F2_texte_uuid)
    t(F1_oeuvre_uuid, lrm("R10_has_member"), F1_texte_uuid)
    t(F2_oeuvre_uuid, lrm("R5_has_component"), F1_texte_uuid)

    # L'oeuvre musicale et composée d'un air
    F1_air_uuid = she(cache.get_uuid([id, "air", "F1", "uuid"], True))
    F2_air_uuid = she(cache.get_uuid([id, "air", "F2", "uuid"], True))
    t(F1_air_uuid, lrm("R3_is_realised_in"), F2_air_uuid)
    t(F1_oeuvre_uuid, lrm("R10_has_member"), F1_air_uuid)
    t(F2_oeuvre_uuid, lrm("R5_has_component"), F2_air_uuid)

    # Note à propos de l'air
    F2_air_note_E13 = she(
        cache.get_uuid([id, "air", "F2", "note sur la musique"], True))
    t(F2_air_note_E13, a, crm("E13_Attribute_Assignement"))
    t(F2_air_note_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
    t(F2_air_note_E13, crm("P140_assigned_attribute_to"), F2_oeuvre_uuid)
    t(F2_air_note_E13, crm("P141_assigned"), l(row[39].value))
    t(F2_air_note_E13, crm("P177_assigned_property_type"), she(""))

    # Rattachement de l'oeuvre musicale à son article
    # TODO cellules multivaluées
    indexation_articles = row[15].value.split("\t")
    for article in indexation_articles:
        id_article = article.lstrip("Mercure Galant/ ").replace(", p. ", "_")
        id_article = id_article.split("-")
        id_article = id_article[0].replace(".", "-")
    id_livraison = id_article.split("_")
    id_livraison = id_livraison[0]

    try:
        ## Livraison originale
        ### Air
        livraison_F2_originale = she(
            cache_corpus.get_uuid(["Corpus", "Livraisons", id_livraison, "Expression originale", "F2"]))
        t(livraison_F2_originale, crm("P148_has_component"), F2_air_uuid)
        ### Texte
        livraison_F2_originale = she(
            cache_corpus.get_uuid(["Corpus", "Livraisons", id_livraison, "Expression originale", "F2"]))
        t(livraison_F2_originale, crm("P148_has_component"), F2_texte_uuid)
        ### Livraison TEI
        ## Air
        livraison_F2_TEI = she(
            cache_corpus.get_uuid(["Corpus", "Livraisons", id_livraison, "Expression TEI", "F2"]))
        t(livraison_F2_TEI, crm("P148_has_component"), F2_air_uuid)
        ## Texte
        livraison_F2_TEI = she(
            cache_corpus.get_uuid(["Corpus", "Livraisons", id_livraison, "Expression TEI", "F2"]))
        t(livraison_F2_TEI, crm("P148_has_component"), F2_texte_uuid)
    except:
        print("L'article ou la livraison", id_article, "(" + id_livraison + ") est introuvable")




#######################################################################################
# CREATION DU GRAPHE ET DU CACHE
#######################################################################################

cache.bye()

serialization = g.serialize(format="turtle", base="http://data-iremus.huma-num.fr/id/")
with open(args.ttl, "wb") as f:
    f.write(serialization)

