# TODO Transformer les auteurs d'ouvrages en E21?

import argparse
from openpyxl import load_workbook
from rdflib import Graph, Namespace, DCTERMS, RDF, RDFS, SKOS, URIRef as u, XSD, Literal as l
import sys
from sherlockcachemanagement import Cache
import requests

# Arguments
parser = argparse.ArgumentParser()
parser.add_argument("--collection_id")
parser.add_argument("--excel_coll")
parser.add_argument("--excel_index")
parser.add_argument("--output_ttl")
parser.add_argument("--cache_images")
parser.add_argument("--cache_corpus")
parser.add_argument("--cache_personnes")
parser.add_argument("--cache_lieux")
parser.add_argument("--cache_vocab_estampes")
args = parser.parse_args()

# Caches
cache_images = Cache(args.cache_images)
cache_corpus = Cache(args.cache_corpus)
cache_personnes = Cache(args.cache_personnes)
cache_lieux = Cache(args.cache_lieux)
cache_vocab_estampes = Cache(args.cache_vocab_estampes)


# Initialisation du graphe
output_graph = Graph()

# Namespaces
crm_ns = Namespace("http://www.cidoc-crm.org/cidoc-crm/")
crmdig_ns = Namespace("http://www.ics.forth.gr/isl/CRMdig/")
iremus_ns = Namespace("http://data-iremus.huma-num.fr/id/")
lrmoo_ns = Namespace("http://www.cidoc-crm.org/lrmoo/")
sdt_ns = Namespace("http://data-iremus.huma-num.fr/datatypes/")
sherlock_ns = Namespace("http://data-iremus.huma-num.fr/ns/sherlock#")

output_graph.bind("crm", crm_ns)
output_graph.bind("dcterms", DCTERMS)
output_graph.bind("lrm", lrmoo_ns)
output_graph.bind("sdt", sdt_ns)
output_graph.bind("skos", SKOS)
output_graph.bind("she", sherlock_ns)
output_graph.bind("crmdig", crmdig_ns)

# Fonctions
a = RDF.type

def crm(x):
	return crm_ns[x]

def crmdig(x):
	return crmdig_ns[x]

def lrm(x):
	return lrmoo_ns[x]

def she(x):
	return iremus_ns[x]

def she_ns(x):
	return sherlock_ns[x]

def t(s, p, o):
	output_graph.add((s, p, o))


# Fichiers Excel
# Index des collections
wb_index = load_workbook(args.excel_index)
index = wb_index.active
# Images de la collection
wb_img = load_workbook(args.excel_coll)
img = wb_img.active

#####################################################################
# LA COLLECTION
#####################################################################

collection_row = None

for row in index:
	if row[0].value == args.collection_id:
		collection_row = row
		break

collection = she(cache_images.get_uuid(["collection", "uuid"], True))
t(collection, a, crmdig("D1_Digital_Object"))
t(collection, RDFS.label, l(collection_row[1].value))
t(collection, crm("P2_has_type"), she("14926d58-83e7-4414-90a8-1a3f5ca8fec1"))
# Creation
collection_E65 = she(cache_images.get_uuid(["collection", "E65"], True))
t(collection_E65, a, crm("E65_Creation"))
t(collection_E65, crm("P94_has_created"), collection)
t(collection_E65, crm("P14_carried_out_by"), she(collection_row[2].value))
# Licence
collection_E30 = she(cache_images.get_uuid(["collection", "E30"], True))
t(collection_E30, a, crm("E30_Right"))
t(collection, crm("P104_is_subject_to"), collection_E30)
t(collection_E30, RDFS.label, l(collection_row[5].value))
# Attribution
t(collection, crm("P105_right_held_by"), she("48a8e9ad-4264-4b0b-a76d-953bc9a34498"))

#####################################################################
# 1. UNE PUBLICATION NUMERISEE
#####################################################################

if collection_row[3].value == "Edition":

	# Work
	livre_F1 = she(cache_images.get_uuid(["collection", "livre", "F1"], True))
	t(livre_F1, a, lrm("F1_Work"))
	livre_E35 = she(cache_images.get_uuid(["collection", "livre", "E41"], True))
	t(livre_F1, crm("P102_has_title"), livre_E35)
	t(livre_E35, a, crm("E35_Title"))
	t(livre_E35, RDFS.label, l(collection_row[1].value))

	# Work Conception
	livre_F27 = she(cache_images.get_uuid(["collection", "livre", "F27"], True))
	t(livre_F27, a, lrm("F27_Work_Conception"))
	t(livre_F27, lrm("R16_initiated"), livre_F1)
	t(livre_F27, crm("P14_carried_out_by"), l(collection_row[7].value))
	if collection_row[8].value != None:
		livre_F27_E52 = she(cache_images.get_uuid(["collection", "livre", "F27", "E52"], True))
		t(livre_F27, crm("P4_has_time-span"), livre_F27_E52)
		t(livre_F27_E52, crm("P80_end_is_qualified_by"), l(collection_row[8].value))

	# Expression
	livre_F2 = she(cache_images.get_uuid(["collection", "livre", "F2"], True))
	t(livre_F1, lrm("R3_is_realised_in"), livre_F2)
	t(livre_F2, a, lrm("F2_Expression"))
	# Expression Creation
	livre_F28 = she(cache_images.get_uuid(["collection", "livre", "F28"], True))
	t(livre_F28, a, lrm("F28_Expression_Creation"))
	t(livre_F28, lrm("R17_created"), livre_F2)
	t(livre_F28, crm("P14_carried_out_by"), l(collection_row[7].value))
	if collection_row[9].value != None:
		livre_F28_E52 = she(cache_images.get_uuid(["collection", "livre", "F28", "E52"], True))
		t(livre_F28, crm("P4_has_time-span"), livre_F28_E52)
		t(livre_F28_E52, crm("P80_end_is_qualified_by"), l(collection_row[9].value))

	# Manifestation
	livre_F3 = she(cache_images.get_uuid(["collection", "livre", "F3"], True))
	t(livre_F3, a, lrm("F3_Manifestation"))
	t(livre_F3, lrm("R4_embodies"), livre_F2)
	## Manifestation Creation
	livre_F30 = she(cache_images.get_uuid(["collection", "livre", "F30"], True))
	t(livre_F30, a, lrm("F30_Manifestation_Creation"))
	t(livre_F30, lrm("R24_created"), livre_F3)
	t(livre_F30, crm("P92_brought_into_existence"), livre_F2)
	if collection_row[10].value != None:
		livre_F30_E52 = she(cache_images.get_uuid(["collection", "livre", "F30", "E52"], True))
		t(livre_F30, crm("P4_has_time-span"), livre_F30_E52)
		t(livre_F30_E52, crm("P80_end_is_qualified_by"), l(collection_row[10].value))
	# Item
	livre_F5 = she(cache_images.get_uuid(["collection", "livre", "F5"], True))
	t(livre_F5, a, lrm("F5_Item"))
	t(livre_F5, lrm("R7_is_materialization_of"), livre_F3)

	#####################################################################
	# LES PAGES DE LA PUBLICATION
	#####################################################################

	img_row = None

	for row in img:
		if row[1].value == args.collection_id:
			img_row = row
			id = img_row[0].value

			# La page comme support physique
			page_E18 = she(cache_images.get_uuid(["collection", "livre", "pages", id, "E18"], True))
			t(page_E18, a, crm("E18_Physical_Object"))
			t(livre_F5, crm("P46_is_composed_of"), page_E18)

			# La page comme support sémiotique
			page_E90 = she(cache_images.get_uuid(["collection", "livre", "pages", id, "E90"], True))
			t(page_E90, a, crm("E90_Symbolic_Object"))
			t(livre_F2, lrm("R15_has_fragment"), page_E90)
			t(page_E18, crm("P128_carries"), page_E90)

			# Identifiant
			page_id = she(cache_images.get_uuid(["collection", "livre", "pages", id, "E42", "id"], True))
			t(page_E90, crm("P1_is_identified_by"), page_id)
			t(page_id, a, crm("E42_Identifier"))
			t(page_id, RDFS.label, l(id))

			# Numéro de la page
			page_no = she(cache_images.get_uuid(["collection", "livre", "pages", id, "E42", "numéro"], True))
			t(page_no, crm("P2_has_type"), she("466bb717-b90f-4104-8f4e-5a13fdde3bc3"))
			t(page_E90, crm("P1_is_identified_by"), page_no)
			t(page_no, a, crm("E42_Identifier"))
			t(page_no, RDFS.label, l(f"Page {img_row[3].value}"))

			# Numérisation de la page
			page_D2 = she(cache_images.get_uuid(["collection", "livre", "pages", "D2"], True))
			t(page_D2, a, crmdig("D2_Digitization_Process"))
			t(page_D2, crmdig("L1_digitized"), page_E18)
			page_D1 = she(cache_images.get_uuid(["collection", "livre", "pages", id, "D1"], True))
			t((page_D1), a, crmdig("D1_Digital_Object"))
			t(page_D2, crmdig("L11_had_output"), page_D1)
			t(page_D1, crm("130_shows_features_of"), page_E90)
			t(collection, crm("P106_is_composed_of"), page_D1)

			# TODO Transcription de la page

#####################################################################
# 2. UNE COLLECTION D'IMAGES INDIVIDUELLES
#####################################################################

def traitement_images(sous_collection):
	for row in img:
		if row[1].value == args.collection_id:
			img_row = row
			id = img_row[0].value

			## E36 Visual Item
			gravure = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "uuid"], True))
			t(gravure, a, crm("E36_Visual_Item"))
			### Identifiant Mercure Galant
			gravure_id_MG = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "Identifiant MG"], True))
			t(gravure_id_MG, a, crm("E42_Identifier"))
			t(gravure_id_MG, crm("P2_has_type"), she("92c258a0-1e34-437f-9686-e24322b95305"))
			t(gravure_id_MG, RDFS.label, l(id))
			t(gravure, crm("P1_is_identified_by"), gravure_id_MG)
			### Identifiant iiif
			gravure_id_iiif = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "Identifiant iiif"], True))
			t(gravure_id_iiif, a, crm("E42_Identifier"))
			t(gravure_id_iiif, crm("P2_has_type"), she("19073c4a-0ef7-4ac4-a51a-e0810a596773"))
			t(gravure_id_iiif, RDFS.label, u("http://data-iremus.huma-num.fr/iiif/mercure-galant/1686-08_320/full/max/0/default.jpg"))
			t(gravure, crm("P1_is_identified_by"), gravure_id_iiif)

			## E12 Production
			if img_row[15].value or img_row[16].value:
				gravure_E12 = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "E12 Production", "uuid"], True))
				t(gravure_E12, a, crm("E12_Production"))
				t(gravure_E12, crm("P108_has_produced"), gravure)

			### Invenit
			if img_row[15].value:
				gravure_invenit = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "E12 Production", "invenit", "uuid"], True))
				t(gravure_invenit, a, crm("E12_Production"))
				t(gravure_invenit, crm("P2_has_type"), she("4d57ac14-247f-4b0e-90ca-0397b6051b8b"))
				t(gravure_E12, crm("P9_consists_of"), gravure_invenit)

				#### E13 Attribute Assignement -  concepteur de la gravure
				gravure_invenit_auteur = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "E12 Production", "invenit", "auteur"], True))
				t(gravure_invenit_auteur, a, crm("E21_Person"))
				t(gravure_invenit_auteur, RDFS.label, l(img_row[15].value))
				gravure_invenit_E13 = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "E12 Production", "invenit", "E13"], True))
				t(gravure_invenit_E13, a, crm("E13_Attribute_Assignement"))
				t(gravure_invenit_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
				t(gravure_invenit_E13, crm("P140_assigned_attribute_to"), gravure_invenit)
				t(gravure_invenit_E13, crm("P141_assigned"), gravure_invenit_auteur)
				t(gravure_invenit_E13, crm("P177_assigned_property_type"), crm("P14_carried_out_by"))

				#### E13 Attribute Assignement - technique de la représentation
				if img_row[12].value:
					gravure_E29 = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "E12 Production", "E29"], True))
					t(gravure_E29, a, crm("E29_Design_or_Procedure"))
					t(gravure_E29, RDFS.label, l(img_row[12].value))
					gravure_E29_E13 = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "E12 Production", "E55", "E13"], True))
					t(gravure_E29_E13, a, crm("E13_Attribute_Assignement"))
					t(gravure_E29_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
					t(gravure_E29_E13, crm("P140_assigned_attribute_to"), gravure_invenit)
					t(gravure_E29_E13, crm("P141_assigned"), gravure_E29)
					t(gravure_E29_E13, crm("P177_assigned_property_type"), crm("P33_used_specific_technique"))

			### Sculpsit
			if img_row[16].value:
				gravure_sculpsit = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "E12 Production", "sculpsit", "uuid"], True))
				t(gravure_sculpsit, a, crm("E12_Production"))
				t(gravure_sculpsit, crm("P2_has_type"), she("f39eb497-5559-486c-b5ce-6a607f615773"))
				t(gravure_E12, crm("P9_consists_of"), gravure_sculpsit)

				#### E13 Attribute Assignement -  sculpteur de la gravure
				gravure_sculpsit_auteur = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "E12 Production", "sculpsit", "auteur"], True))
				t(gravure_sculpsit_auteur, a, crm("E21_Person"))
				t(gravure_sculpsit_auteur, RDFS.label, l(img_row[16].value))
				gravure_sculpsit_E13 = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "E12 Production", "sculpsit", "E13"], True))
				t(gravure_sculpsit_E13, a, crm("E13_Attribute_Assignement"))
				t(gravure_sculpsit_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
				t(gravure_sculpsit_E13, crm("P140_assigned_attribute_to"), gravure_sculpsit)
				t(gravure_sculpsit_E13, crm("P141_assigned"), gravure_sculpsit_auteur)
				t(gravure_sculpsit_E13, crm("P177_assigned_property_type"), crm("P14_carried_out_by"))

				#### E13 Attribute Assignement - technique de la gravure
				if img_row[13].value:
					gravure_E55 = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "E12 Production", "E55", "uuid"], True))
					t(gravure_E55, a, crm("E55_Type"))
					t(gravure_E55, RDFS.label, l(img_row[13].value))
					gravure_E55_E13 = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "E12 Production", "E55", "E13"], True))
					t(gravure_E55_E13, a, crm("E13_Attribute_Assignement"))
					t(gravure_E55_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
					t(gravure_E55_E13, crm("P140_assigned_attribute_to"), gravure_sculpsit)
					t(gravure_E55_E13, crm("P141_assigned"), gravure_E55)
					t(gravure_E55_E13, crm("P177_assigned_property_type"), crm("P32_used_general_technique"))

			## Rattachement à la livraison ou à l'article
			### Si l'article n'est pas précisé:
			if not img_row[3].value :
				id_image = f"MG-{id}"
				id_livraison = f"MG-{id[0:-4]}"
				if id_livraison.endswith("_"):
					id_livraison = id_livraison[0:-1]
				try:
					#### Livraison originale
					livraison_F2_originale = she(
						cache_corpus.get_uuid(["Corpus", "Livraisons", id_livraison, "Expression originale", "F2"]))
					t(livraison_F2_originale, crm("P148_has_component"), gravure)
					#### Livraison TEI
					livraison_F2_TEI = she(
						cache_corpus.get_uuid(["Corpus", "Livraisons", id_livraison, "Expression TEI", "F2"]))
					t(livraison_F2_TEI, crm("P148_has_component"), gravure)
				except:
					print("L'image " + id_image + " n'est reliée à aucune livraison")

			### Si l'article est précisé:
			else:
				id_article = img_row[3].value
				id_livraison = id_article[0:11]
				try:
					if id_livraison.endswith("_"):
						id_livraison = id_livraison[0:-1]
						article_F2 = she(cache_corpus.get_uuid(
							["Corpus", "Livraisons", id_livraison, "Expression originale", "Articles", id_article, "F2"]))
						t(article_F2, crm("P148_has_component"), gravure)
					else:
						article_F2 = she(cache_corpus.get_uuid(
							["Corpus", "Livraisons", id_livraison, "Expression originale", "Articles", id_article, "F2"]))
						t(article_F2, crm("P148_has_component"), gravure)
				except:
					print("Article contenant la gravure : l'article " + id_article + " est introuvable dans les fichiers TEI")

			## Article annexe à la gravure
			if img_row[4].value:
				id_article = img_row[4].value
				id_livraison = id_article[0:10]
				try:
					uuid_article = she(cache_corpus.get_uuid(
						["Corpus", "Livraisons", id_livraison, "Expression originale", "Articles", id_article,
						 "F2"]))
					gravure_seeAlso_E13 = she(
						cache_images.get_uuid(["collection", id, "gravure (E36)", "seeAlso", "E13"], True))
					t(gravure_seeAlso_E13, a, crm("E13_Attribute_Assignement"))
					t(gravure_seeAlso_E13, crm("P14_carried_out_by"),
					  she("684b4c1a-be76-474c-810e-0f5984b47921"))
					t(gravure_seeAlso_E13, crm("P140_assigned_attribute_to"), gravure)
					t(gravure_seeAlso_E13, crm("P141_assigned"), uuid_article)
					t(gravure_seeAlso_E13, crm("P177_assigned_property_type"), RDFS.seeAlso)
					### Commentaire décrivant le lien entre la gravure et l'article
					if img_row[5].value:
						gravure_seeAlso_P3_E13 = she(
							cache_images.get_uuid(["collection", id, "gravure (E36)", "seeAlso", "note", "E13"], True))
						t(gravure_seeAlso_P3_E13, a, crm("E13_Attribute_Assignement"))
						t(gravure_seeAlso_P3_E13, crm("P14_carried_out_by"),
						  she("684b4c1a-be76-474c-810e-0f5984b47921"))
						t(gravure_seeAlso_P3_E13, crm("P140_assigned_attribute_to"), uuid_article)
						t(gravure_seeAlso_P3_E13, crm("P141_assigned"), l(img_row[5].value))
						t(gravure_seeAlso_P3_E13, crm("P177_assigned_property_type"), crm("P3_has_note"))
				except:
					print(
						"Article annexe à la gravure : l'article " + id_article + " est introuvable dans les fichiers TEI")

			## Lien Gallica
			if img_row[6].value:
				lien_gallica = u(img_row[6].value)
				t(lien_gallica, crm("P2_has_type"), she("e73699b0-9638-4a9a-bfdd-ed1715416f02"))
				img_gallica_E13 = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "gallica", "E13"], True))
				t(img_gallica_E13, a, crm("E13_Attribute_Assignement"))
				t(img_gallica_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
				t(img_gallica_E13, crm("P140_assigned_attribute_to"), gravure)
				t(img_gallica_E13, crm("P141_assigned"), lien_gallica)
				t(img_gallica_E13, crm("P177_assigned_property_type"), RDFS.seeAlso)

			## Titre sur l'image (E13)
			if img_row[7].value:
				gravure_titre = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "titre sur l'image"], True))
				t(gravure_titre, a, crm("E13_Attribute_Assignement"))
				t(gravure_titre, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
				t(gravure_titre, crm("P140_assigned_attribute_to"), gravure)
				t(gravure_titre, crm("P141_assigned"), l(img_row[7].value))
				t(gravure_titre, crm("P177_assigned_property_type"), she("01a07474-f2b9-4afd-bb05-80842ecfb527"))

			## Titre descriptif/forgé (E13)
			if img_row[8].value:
				gravure_titre = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "titre forgé"], True))
				t(gravure_titre, a, crm("E13_Attribute_Assignement"))
				t(gravure_titre, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
				t(gravure_titre, crm("P140_assigned_attribute_to"), gravure)
				t(gravure_titre, crm("P141_assigned"), l(img_row[8].value))
				t(gravure_titre, crm("P177_assigned_property_type"), she("58fb99dd-1ffb-4e00-a16f-ef6898902301"))

			## Titre dans le péritexte (E13)
			if img_row[9].value:
				gravure_titre = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "titre péritexte"], True))
				t(gravure_titre, a, crm("E13_Attribute_Assignement"))
				t(gravure_titre, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
				t(gravure_titre, crm("P140_assigned_attribute_to"), gravure)
				t(gravure_titre, crm("P141_assigned"), l(img_row[9].value))
				t(gravure_titre, crm("P177_assigned_property_type"), she("ded9ea93-b400-4550-9aa8-e5aac1d627a0"))

			## Lieu représenté
			if img_row[14].value:

				lieu = img_row[14].value

				### Zone de l'image comportant la représentation du lieu (E13)
				gravure_zone_img = she(cache_images.get_uuid(["collection", id, "gravure (E36)", lieu, "zone de l'image (E36)", "uuid"], True))
				t(gravure_zone_img, a, crm("E36_Visual_Item"))
				gravure_zone_img_E13 = she(cache_images.get_uuid(["collection", id, "gravure (E36)", lieu, "zone de l'image (E36)", "E13"], True))
				t(gravure_zone_img_E13, a, crm("E13_Attribute_Assignement"))
				t(gravure_zone_img_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
				t(gravure_zone_img_E13, crm("P140_assigned_attribute_to"), gravure)
				t(gravure_zone_img_E13, crm("P141_assigned"), gravure_zone_img)
				t(gravure_zone_img_E13, crm("P177_assigned_property_type"), crm("P106_is_composed_of"))

				### Recherche d'UUID dans le référentiel des lieux
				try:
					lieu_uuid = she(cache_lieux.get_uuid(["lieu", str(lieu), "E93", "uuid"]))
					if lieu_uuid:
						gravure_lieu_E13 = she(cache_images.get_uuid(["collection", id, "gravure (E36)", lieu, "zone de l'image (E36)", "lieu représenté"], True))
						t(gravure_lieu_E13, a, crm("E13_Attribute_Assignement"))
						t(gravure_lieu_E13, crm("P14_carried_out_by"),
						  she("684b4c1a-be76-474c-810e-0f5984b47921"))
						t(gravure_lieu_E13, crm("P140_assigned_attribute_to"), gravure_zone_img)
						t(gravure_lieu_E13, crm("P177_assigned_property_type"), crm("P138_represents"))
						t(gravure_lieu_E13, crm("P141_assigned"), lieu_uuid)

				except:
					gravure_lieu_E13 = she(cache_images.get_uuid(
						["collection", id, "gravure (E36)", lieu, "zone de l'image (E36)", "lieu représenté"], True))
					t(gravure_lieu_E13, a, crm("E13_Attribute_Assignement"))
					t(gravure_lieu_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
					t(gravure_lieu_E13, crm("P140_assigned_attribute_to"), gravure_zone_img)
					t(gravure_lieu_E13, crm("P177_assigned_property_type"), crm("P138_represents"))
					t(gravure_lieu_E13, crm("P141_assigned"), l(lieu))

			## Objet/Personne représentée (E13)
			if img_row[11].value:

				sujets = img_row[11].value.split(";")

				for sujet in sujets:
					sujet = sujet.strip()

					### Zone de l'image comportant la représentation du sujet (E13)
					gravure_zone_img = she(
						cache_images.get_uuid(["collection", id, "gravure (E36)", sujet, "zone de l'image (E36)", "uuid"], True))
					t(gravure_zone_img, a, crm("E36_Visual_Item"))
					gravure_zone_img_E13 = she(
						cache_images.get_uuid(["collection", id, "gravure (E36)", sujet, "zone de l'image (E36)", "E13"], True))
					t(gravure_zone_img_E13, a, crm("E13_Attribute_Assignement"))
					t(gravure_zone_img_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
					t(gravure_zone_img_E13, crm("P140_assigned_attribute_to"), gravure)
					t(gravure_zone_img_E13, crm("P141_assigned"), gravure_zone_img)
					t(gravure_zone_img_E13, crm("P177_assigned_property_type"), crm("P106_is_composed_of"))

					### Recherche d'UUID dans le référentiel des personnes
					try:
						personne_uuid = she(cache_personnes.get_uuid(["personnes", sujet, "uuid"]))
						if personne_uuid:
							gravure_personne_E13 = she(cache_images.get_uuid(
								["collection", id, "gravure (E36)", sujet, "zone de l'image (E36)", "personne représentée"],
								True))
							t(gravure_personne_E13, a, crm("E13_Attribute_Assignement"))
							t(gravure_personne_E13, crm("P14_carried_out_by"),
							  she("684b4c1a-be76-474c-810e-0f5984b47921"))
							t(gravure_personne_E13, crm("P140_assigned_attribute_to"), gravure_zone_img)
							t(gravure_personne_E13, crm("P177_assigned_property_type"), crm("P138_represents"))
							t(gravure_personne_E13, crm("P141_assigned"), personne_uuid)
					except:

				### Recherche d'UUID dans le vocabulaire d'indexation des gravures
						try:
							objet_uuid = she(cache_vocab_estampes.get_uuid(["vocabulaire indexation gravures", sujet, "uuid"]))
							if objet_uuid:

								if "médaille" not in sujet:
									gravure_objet_E13 = she(cache_images.get_uuid(
										["collection", id, "gravure (E36)", sujet, "zone de l'image (E36)", "objet représenté"], True))
									t(gravure_objet_E13, a, crm("E13_Attribute_Assignement"))
									t(gravure_objet_E13, crm("P14_carried_out_by"),
									  she("684b4c1a-be76-474c-810e-0f5984b47921"))
									t(gravure_objet_E13, crm("P140_assigned_attribute_to"), gravure_zone_img)
									t(gravure_objet_E13, crm("P177_assigned_property_type"), crm("P138_represents"))
									t(gravure_objet_E13, crm("P141_assigned"), objet_uuid)

								### Si l'objet représenté est une médaille (E13)
								if "médaille" in sujet:
									#### E13 Attribute Assignement - la médaille
									t(objet_uuid, a, crm("E55_Type"))
									t(objet_uuid, crm("P2_has_type"), she("4b51d9dc-3623-47f4-ab45-239604e18930"))

									gravure_médaille_E13 = she(cache_images.get_uuid(
										["collection", id, "gravure (E36)", sujet, "zone de l'image (E36)", "objet représenté"], True))
									t(gravure_médaille_E13, a, crm("E13_Attribute_Assignement"))
									t(gravure_médaille_E13, crm("P14_carried_out_by"),
									  she("684b4c1a-be76-474c-810e-0f5984b47921"))
									t(gravure_médaille_E13, crm("P140_assigned_attribute_to"), gravure_zone_img)
									t(gravure_médaille_E13, crm("P177_assigned_property_type"), crm("P138_represents"))
									t(gravure_médaille_E13, crm("P141_assigned"), objet_uuid)

									#### Si la médaille comporte une inscription
									if img_row[17].value or img_row[18].value:
										gravure_médaille_inscrip_E36 = she(
											cache_images.get_uuid(
												["collection", id, "gravure (E36)", sujet, "zone d'inscription", "uuid"], True))
										t(gravure_médaille_inscrip_E36, a, crm("E36_Visual_Item"))
										t(gravure_médaille_inscrip_E36, RDFS.label, l("Zone d'inscription"))
										##### E13 Attribute Assignement
										gravure_médaille_inscrip_E13 = she(
											cache_images.get_uuid(
												["collection", id, "gravure (E36)", sujet, "zone d'inscription", "E13"], True))
										t(gravure_médaille_inscrip_E13, a, crm("E13_Attribute_Assignement"))
										t(gravure_médaille_inscrip_E13, crm("P14_carried_out_by"),
										  she("684b4c1a-be76-474c-810e-0f5984b47921"))
										t(gravure_médaille_inscrip_E13, crm("P140_assigned_attribute_to"), objet_uuid)
										t(gravure_médaille_inscrip_E13, crm("P141_assigned"), gravure_médaille_inscrip_E36)
										t(gravure_médaille_inscrip_E13, crm("P177_assigned_property_type"),
										  crm("P106_is_composed_of"))

									##### Si la médaille comporte une inscription en légende (E13)
									if img_row[17].value:
										gravure_médaille_inscrip_E33 = she(
											cache_images.get_uuid(
												["collection", id, "gravure (E36)", sujet, "zone d'inscription",
												 "inscription", "en légende", "uuid"], True))
										t(gravure_médaille_inscrip_E33, a, crm("E33_Linguistic_Object"))

										##### E13 Attribute Assignement du E33
										gravure_médaille_inscrip_E33_E13 = she(
											cache_images.get_uuid(
												["collection", id, "gravure (E36)", sujet, "zone d'inscription",
												 "inscription", "en légende", "E13"], True))
										t(gravure_médaille_inscrip_E33_E13, a, crm("E13_Attribute_Assignement"))
										t(gravure_médaille_inscrip_E33_E13, crm("P14_carried_out_by"),
										  she("684b4c1a-be76-474c-810e-0f5984b47921"))
										t(gravure_médaille_inscrip_E33_E13, crm("P140_assigned_attribute_to"),
										  gravure_médaille_inscrip_E36)
										t(gravure_médaille_inscrip_E33_E13, crm("P141_assigned"), gravure_médaille_inscrip_E33)
										t(gravure_médaille_inscrip_E33_E13, crm("P177_assigned_property_type"),
										  crm("P165_incorporates"))
										t(gravure_médaille_inscrip_E33_E13,
										  she("sheP_position_du_texte_par_rapport_à_la_médaille"),
										  she("fc229531-0999-4499-ab0b-b45e18e8196f"))

										##### E13 Attribute Assignement du contenu de l'inscription
										gravure_médaille_inscrip_P190_E13 = she(
											cache_images.get_uuid(
												["collection", id, "gravure (E36)", sujet, "zone d'inscription",
												 "inscription", "en légende", "contenu"], True))
										t(gravure_médaille_inscrip_P190_E13, a, crm("E13_Attribute_Assignement"))
										t(gravure_médaille_inscrip_P190_E13, crm("P14_carried_out_by"),
										  she("684b4c1a-be76-474c-810e-0f5984b47921"))
										t(gravure_médaille_inscrip_P190_E13, crm("P140_assigned_attribute_to"),
										  gravure_médaille_inscrip_E33)
										t(gravure_médaille_inscrip_P190_E13, crm("P141_assigned"), l(img_row[17].value))
										t(gravure_médaille_inscrip_P190_E13, crm("P177_assigned_property_type"),
										  crm("P190_has_symbolic_content"))

									##### Si la médaille comporte une inscription en exergue (E13)
									if img_row[18].value:
										gravure_médaille_inscrip_E33 = she(
											cache_images.get_uuid(
												["collection", id, "gravure (E36)", sujet, "zone d'inscription",
												 "inscription", "en exergue", "uuid"], True))
										t(gravure_médaille_inscrip_E33, a, crm("E33_Linguistic_Object"))

										##### E13 Attribute Assignement du E33
										gravure_médaille_inscrip_E33_E13 = she(
											cache_images.get_uuid(
												["collection", id, "gravure (E36)", sujet, "zone d'inscription",
												 "inscription", "en exergue", "E13"], True))
										t(gravure_médaille_inscrip_E33_E13, a, crm("E13_Attribute_Assignement"))
										t(gravure_médaille_inscrip_E33_E13, crm("P14_carried_out_by"),
										  she("684b4c1a-be76-474c-810e-0f5984b47921"))
										t(gravure_médaille_inscrip_E33_E13, crm("P140_assigned_attribute_to"),
										  gravure_médaille_inscrip_E36)
										t(gravure_médaille_inscrip_E33_E13, crm("P141_assigned"), gravure_médaille_inscrip_E33)
										t(gravure_médaille_inscrip_E33_E13, crm("P177_assigned_property_type"),
										  crm("P165_incorporates"))
										t(gravure_médaille_inscrip_E33_E13,
										  she("sheP_position_du_texte_par_rapport_à_la_médaille"),
										  she("357a459f-4f27-4d46-b5ac-709a410bce04"))

										##### E13 Attribute Assignement du contenu de l'inscription
										gravure_médaille_inscrip_P190_E13 = she(
											cache_images.get_uuid(
												["collection", id, "gravure (E36)", sujet, "zone d'inscription",
												 "inscription", "en exergue", "contenu"], True))
										t(gravure_médaille_inscrip_P190_E13, a, crm("E13_Attribute_Assignement"))
										t(gravure_médaille_inscrip_P190_E13, crm("P14_carried_out_by"),
										  she("684b4c1a-be76-474c-810e-0f5984b47921"))
										t(gravure_médaille_inscrip_P190_E13, crm("P140_assigned_attribute_to"),
										  gravure_médaille_inscrip_E33)
										t(gravure_médaille_inscrip_P190_E13, crm("P141_assigned"), l(img_row[18].value))
										t(gravure_médaille_inscrip_P190_E13, crm("P177_assigned_property_type"),
										  crm("P190_has_symbolic_content"))

						except:
							gravure_objet_E13 = she(cache_images.get_uuid(
								["collection", id, "gravure (E36)", sujet, "zone de l'image (E36)", "objet représenté"], True))
							t(gravure_objet_E13, a, crm("E13_Attribute_Assignement"))
							t(gravure_objet_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
							t(gravure_objet_E13, crm("P140_assigned_attribute_to"), gravure_zone_img)
							t(gravure_objet_E13, crm("P177_assigned_property_type"), crm("P138_represents"))
							t(gravure_objet_E13, crm("P141_assigned"), l(sujet))


			## Notes sur la provenance de la gravure
			if img_row[19].value:
				gravure_notes_E13 = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "notes", "E13"], True))
				t(gravure_notes_E13, a, crm("E13_Attribute_Assignement"))
				t(gravure_notes_E13, crm("P14_carried_out_by"), she("684b4c1a-be76-474c-810e-0f5984b47921"))
				t(gravure_notes_E13, crm("P140_assigned_attribute_to"), gravure)
				t(gravure_notes_E13, crm("P141_assigned"), l(img_row[19].value))
				t(gravure_notes_E13, crm("P177_assigned_property_type"), crm("P3_has_note"))

			## Autres liens externes
			if img_row[20].value:
				try:
					response = requests.get(img_row[20].value)
					lien_externe = u(img_row[20].value)
					gravure_lien_externe_E13 = she(
						cache_images.get_uuid(["collection", id, "gravure (E36)", "lien externe", "E13"], True))
					t(gravure_lien_externe_E13, a, crm("E13_Attribute_Assignement"))
					t(gravure_lien_externe_E13, crm("P14_carried_out_by"),
					  she("684b4c1a-be76-474c-810e-0f5984b47921"))
					t(gravure_lien_externe_E13, crm("P140_assigned_attribute_to"), gravure)
					t(gravure_lien_externe_E13, crm("P141_assigned"), lien_externe)
					t(gravure_lien_externe_E13, crm("P177_assigned_property_type"), RDFS.seeAlso)
				except:
					print("'" + img_row[20].value + "' n'est pas une URL valide")


			## Bibliographie relative à la gravure
			if img_row[21].value:
				biblio = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "bibliographie", "uuid"], True))
				t(biblio, a, crm("E31_Document"))
				t(biblio, RDFS.label, l(img_row[21].value))
				## E13 Attribute Assignement
				gravure_biblio_E13 = she(cache_images.get_uuid(["collection", id, "gravure (E36)", "bibliographie", "E13"], True))
				t(gravure_biblio_E13, a, crm("E13_Attribute_Assignement"))
				t(gravure_biblio_E13, crm("P14_carried_out_by"),
				  she("684b4c1a-be76-474c-810e-0f5984b47921"))
				t(gravure_biblio_E13, crm("P140_assigned_attribute_to"), gravure)
				t(gravure_biblio_E13, crm("P141_assigned"), biblio)
				t(gravure_biblio_E13, crm("P177_assigned_property_type"), crm("P70_documents"))

#####################################################################
# CHOIX DE LA COLLECTION A TRAITER
#####################################################################

if collection_row[3].value == "Images":

	img_row = None

	if args.collection_id == "mg_estampes":
		coll_estampes = she(cache_images.get_uuid(["collection", "estampes"], True))
		print("Je me trouve dans la collection d'estampes")
		t(coll_estampes, a, crmdig("D1_Digital_Object"))
		t(coll_estampes, RDFS.label, l("Collection d'estampes"))
		t(coll_estampes, crm("P2_has_type"), she("20674932-def8-4a73-9b67-ac49a16b5243"))
		t(collection, crm("P106_is_composed_of"), coll_estampes)
		traitement_images(coll_estampes)

	if args.collection_id == "MGM":
		coll_musique = she(cache_images.get_uuid(["collection", "musique"], True))
		print("Je me trouve dans la collection de partitions")
		t(coll_musique, a, crmdig("D1_Digital_Object"))
		t(coll_musique, RDFS.label, l("Collection de partitions"))
		t(coll_musique, crm("P2_has_type"), she("fa4959c8-8537-4ef0-8d58-4f761e9679b1"))
		t(collection, crm("P106_is_composed_of"), coll_musique)
		traitement_images(coll_musique)



serialization = output_graph.serialize(format="turtle", base="http://data-iremus.huma-num.fr/id/")
with open(args.output_ttl, "wb") as f:
	f.write(serialization)

cache_images.bye()
