# https://gist.github.com/adejones/e0d8783dad234ed400420e0193b692a7

from openpyxl import load_workbook


class Ikselesix:
    def __init__(self, path):
        self.book = load_workbook(filename=path)
        self._dict = {}
        for sheetname in self.book.sheetnames:
            sheet = self.book[sheetname]
            rows = sheet.max_row
            cols = sheet.max_column

            def item(i, j):
                return (sheet.cell(row=1, column=j).value, sheet.cell(row=i, column=j).value)

            sheet_rows = []
            for r in range(2, rows + 1):
                row = {}
                for c in range(1, cols + 1):
                    i = item(r, c)
                    if i[0]:
                        row[i[0]] = i[1]
                if len(set(row.values())) == 1 and next(iter(set(row.values()))) == None:
                    continue
                sheet_rows.append(row)
            self._dict[sheetname] = sheet_rows

    def __getitem__(self, sheetname):
        return self._dict[sheetname]
